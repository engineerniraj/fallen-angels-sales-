import hmac
import io
import re
import zipfile
from copy import copy

import google.generativeai as genai
import openpyxl
import streamlit as st
from PIL import Image
from pdf2image import convert_from_bytes
from pillow_heif import register_heif_opener

register_heif_opener()

st.set_page_config(page_title="Niraj Excel Tools", page_icon="📊", layout="wide")

# ─── Auth ────────────────────────────────────────────────────────────────────

def check_password():
    if st.session_state.get("authenticated"):
        return True
    st.title("Niraj Excel Tools")
    st.subheader("Login required")
    username = st.text_input("Username")
    password = st.text_input("Password", type="password")
    if st.button("Login", type="primary"):
        valid_username = st.secrets.get("APP_USERNAME", "")
        valid_password = st.secrets.get("APP_PASSWORD", "")
        if hmac.compare_digest(username, valid_username) and hmac.compare_digest(password, valid_password):
            st.session_state["authenticated"] = True
            st.rerun()
        else:
            st.error("Incorrect username or password")
    return False


def logout_button():
    with st.sidebar:
        if st.button("Logout"):
            st.session_state["authenticated"] = False
            st.rerun()


# ─── Constants ───────────────────────────────────────────────────────────────

# Maps show name → sheet index (0-based) and the exact day label in the Excel
SHOWS = [
    "Tuesday",
    "Wednesday 2 PM",
    "Wednesday 8 PM",
    "Thursday",
    "Friday",
    "Saturday 2 PM",
    "Saturday 8 PM",
    "Sunday",
]

# Exact product labels used in the review table and Gemini prompt
PRODUCTS = [
    "Poster",
    "Magnet",
    "Lapel Pin",
    "Keychain",
    "Mug",
    "Tote",
    "Logo Tee S",
    "Logo Tee M",
    "Logo Tee L",
    "Logo Tee XL",
    "Logo Tee XXL",   # added – Logo tee also has XXL in the invoice sheet
    "Lapse Tee S",
    "Lapse Tee M",
    "Lapse Tee L",
    "Lapse Tee XL",
    "Lapse Tee XXL",
    "Hoodie S",
    "Hoodie M",
    "Hoodie L",
    "Hoodie XL",
    "Hoodie XXL",     # added – Hoodie also has XXL in the invoice sheet
]

# Hard-coded row map: product label → Excel row number (same across all day sheets)
# Column I (index 9) = "Retail IN" for Position #1
PRODUCT_ROW_MAP = {
    "Poster":        24,
    "Magnet":        25,
    "Lapel Pin":     26,
    "Keychain":      28,
    "Mug":           29,
    "Tote":          30,
    # Logo Fitted Tee sizes
    "Logo Tee S":    73,
    "Logo Tee M":    74,
    "Logo Tee L":    75,
    "Logo Tee XL":   76,
    "Logo Tee XXL":  77,
    # Lapse Tee sizes (skip row 78 which is the 'Y'/youth row)
    "Lapse Tee S":   79,
    "Lapse Tee M":   80,
    "Lapse Tee L":   81,
    "Lapse Tee XL":  82,
    "Lapse Tee XXL": 83,
    # Hoodie sizes (skip row 85 which is actually S)
    "Hoodie S":      85,
    "Hoodie M":      86,
    "Hoodie L":      87,
    "Hoodie XL":     88,
    "Hoodie XXL":    89,
}

COLUMN_I = 9   # Column I = retail quantity sold (Position #1 IN)


# ─── Excel writing ───────────────────────────────────────────────────────────

def copy_cell_style(source_cell, target_cell):
    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.fill = copy(source_cell.fill)
        target_cell.border = copy(source_cell.border)
        target_cell.alignment = copy(source_cell.alignment)
        target_cell.number_format = source_cell.number_format
        target_cell.protection = copy(source_cell.protection)


def write_retail_values(workbook, sheet_entries):
    """
    Write retail quantities into Column I of each day sheet.
    sheet_entries: { show_name: { product_label: qty } }
    Returns a summary list.
    """
    summary = []
    for sheet_idx, show_name in enumerate(SHOWS):
        if sheet_idx >= len(workbook.worksheets):
            continue
        ws = workbook.worksheets[sheet_idx]
        entries = sheet_entries.get(show_name, {})
        entered, missing = 0, []

        for product in PRODUCTS:
            qty = int(entries.get(product, 0) or 0)
            if qty == 0:
                continue
            row = PRODUCT_ROW_MAP.get(product)
            if row is None:
                missing.append(product)
                continue
            cell = ws.cell(row=row, column=COLUMN_I)
            # Copy style from adjacent column for consistency
            copy_cell_style(ws.cell(row=row, column=COLUMN_I + 1), cell)
            cell.value = qty
            entered += 1

        summary.append({
            "show": show_name,
            "worksheet": ws.title,
            "items_written": entered,
            "missing_products": missing,
        })
    return summary


def build_download(master_file_bytes, sheet_entries):
    workbook = openpyxl.load_workbook(io.BytesIO(master_file_bytes))
    summary = write_retail_values(workbook, sheet_entries)
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output, summary


# ─── Gemini OCR ──────────────────────────────────────────────────────────────

def get_gemini_model():
    api_key = st.secrets.get("GEMINI_API_KEY", "")
    if not api_key:
        raise ValueError("GEMINI_API_KEY is missing in Streamlit Secrets.")
    genai.configure(api_key=api_key)
    return genai.GenerativeModel("gemini-2.5-flash")


def build_gemini_prompt():
    # Build the exact expected output block so Gemini knows the format
    example_lines = "\n".join(f"{p}|0" for p in PRODUCTS)
    return f"""You are reading a handwritten FALLEN ANGELS merchandise sales sheet image.

=== SHEET COLUMN LAYOUT (left to right) ===
ITEM | SIZE | RETAIL_PRICE | STAFF_PRICE | PRE | ADD | START | END | PROMO | RETAIL_QTY | STAFF_QTY | $RETAIL | $STAFF

=== YOUR ONLY JOB ===
For each product row, find HOW MANY UNITS were SOLD at RETAIL.

=== HOW TO FIND THE SOLD QTY ===
Step 1: Look at the RETAIL_QTY column (the column RIGHT AFTER the PROMO column).
        - If it has a handwritten number → that IS the sold qty. Use it directly.
Step 2: If RETAIL_QTY is blank or empty:
        - Calculate: sold = START - END
        - If START = END → sold = 0
        - If START or END is unreadable → sold = 0
Step 3: NEVER use the $RETAIL column (dollar amounts like $25, $180) — those are dollar totals, not qty.
Step 4: NEVER use PRE or ADD columns for qty.

=== PRODUCTS TO EXTRACT ===
The invoice sheet has these sections IN THIS ORDER:
1. POSTER (single row)
2. MAGNET (single row)
3. LAPEL PIN (single row)
4. KEYCHAIN (single row)
5. MUG (single row)
6. TOTE (single row)
7. LOGO (FITTED) tee — rows for S, M, L, XL, XXL
8. LAPSE TEE (UNISEX) — rows for S, M, L, XL, XXL
9. HOODIE — rows for S, M, L, XL, XXL

=== SIZE MATCHING RULES ===
- Each size (S, M, L, XL, XXL) is a SEPARATE row. Read each row individually.
- S and XL look similar in handwriting — check carefully.
- XXL and XL are DIFFERENT. XXL has two X's.
- Logo Tee and Lapse Tee are COMPLETELY SEPARATE sections. Do not mix them.
- Hoodie is separate from all tees.

=== OUTPUT FORMAT ===
Return EXACTLY {len(PRODUCTS)} lines. No extra text, no markdown, no explanations.
Format: PRODUCT_NAME|NUMBER

Expected output (fill in correct numbers):
{example_lines}"""


def images_from_upload(uploaded_file):
    file_bytes = uploaded_file.getvalue()
    name = uploaded_file.name.lower()
    if name.endswith(".pdf"):
        pages = convert_from_bytes(file_bytes, dpi=300)
        return pages
    img = Image.open(io.BytesIO(file_bytes))
    # Upscale small images for better OCR
    if img.width < 1500:
        scale = 1500 / img.width
        img = img.resize((int(img.width * scale), int(img.height * scale)), Image.LANCZOS)
    return [img]


def extract_entries_from_invoice(uploaded_file, model):
    """Run Gemini OCR on one invoice file. Returns dict {product: qty} and raw text."""
    pages = images_from_upload(uploaded_file)
    prompt = build_gemini_prompt()
    all_text = []
    for page in pages:
        img = page.convert("RGB")
        try:
            response = model.generate_content([prompt, img])
            text = response.text or ""
            all_text.append(text)
        except Exception as e:
            all_text.append(f"ERROR: {e}")
    combined = "\n".join(all_text)
    return parse_gemini_output(combined), combined


def parse_gemini_output(text):
    """Parse Gemini's PRODUCT|QTY output into a dict. Robust to minor formatting issues."""
    entries = {p: 0 for p in PRODUCTS}
    normalized_map = {p.strip().lower(): p for p in PRODUCTS}

    for line in text.splitlines():
        line = line.strip()
        # Remove any markdown artifacts
        line = line.lstrip("- *#").strip()
        if "|" not in line:
            continue
        parts = line.split("|", 1)
        raw_label = parts[0].strip().rstrip(":")
        raw_qty = parts[1].strip() if len(parts) > 1 else "0"

        label = normalized_map.get(raw_label.lower())
        if not label:
            # Try partial match as fallback
            for key, canonical in normalized_map.items():
                if raw_label.lower() in key or key in raw_label.lower():
                    label = canonical
                    break
        if not label:
            continue

        # Extract the number — take the first reasonable number (not 0 if others exist)
        numbers = [int(n) for n in re.findall(r"\b\d+\b", raw_qty)]
        if numbers:
            # Sanity check: qty sold per show is unlikely to exceed 50
            qty = numbers[0] if numbers[0] <= 50 else 0
            entries[label] = qty

    return entries


# ─── Show detection ──────────────────────────────────────────────────────────

def detect_show_from_filename(filename):
    name = filename.lower()
    if "tuesday" in name:
        return "Tuesday"
    if "wednesday" in name and any(x in name for x in ["2pm", "2_pm", "2 pm", "_2_", "_2."]):
        return "Wednesday 2 PM"
    if "wednesday" in name and any(x in name for x in ["8pm", "8_pm", "8 pm", "_8_", "_8."]):
        return "Wednesday 8 PM"
    if "wednesday" in name:
        # fallback: if only one wednesday, try to pick from 2/8
        if "2" in name:
            return "Wednesday 2 PM"
        if "8" in name:
            return "Wednesday 8 PM"
    if "thursday" in name:
        return "Thursday"
    if "friday" in name:
        return "Friday"
    if "saturday" in name and any(x in name for x in ["2pm", "2_pm", "2 pm", "_2_", "2."]):
        return "Saturday 2 PM"
    if "saturday" in name and any(x in name for x in ["8pm", "8_pm", "8 pm", "_8_", "8."]):
        return "Saturday 8 PM"
    if "saturday" in name:
        if "2" in name:
            return "Saturday 2 PM"
        if "8" in name:
            return "Saturday 8 PM"
    if "sunday" in name:
        return "Sunday"
    return None


# ─── UI helpers ──────────────────────────────────────────────────────────────

def entries_to_review_rows(sheet_entries):
    rows = []
    for show in SHOWS:
        entries = sheet_entries.get(show, {})
        row = {"Show": show}
        for p in PRODUCTS:
            row[p] = int(entries.get(p, 0) or 0)
        rows.append(row)
    return rows


def review_rows_to_sheet_entries(rows):
    """Convert data_editor output (list of dicts OR DataFrame) to sheet_entries dict."""
    sheet_entries = {show: {p: 0 for p in PRODUCTS} for show in SHOWS}

    # Handle both DataFrame and list-of-dicts (Streamlit returns DataFrame in newer versions)
    try:
        import pandas as pd
        if isinstance(rows, pd.DataFrame):
            rows = rows.to_dict("records")
    except ImportError:
        pass

    for row in rows:
        show = row.get("Show")
        if show not in sheet_entries:
            continue
        for p in PRODUCTS:
            try:
                val = row.get(p, 0)
                sheet_entries[show][p] = int(val) if val not in (None, "", "None") else 0
            except Exception:
                sheet_entries[show][p] = 0
    return sheet_entries


# ─── Pages ───────────────────────────────────────────────────────────────────

def automated_ocr_processor():
    st.header("Automated Invoice OCR → Excel")
    st.caption("Upload the master Excel + invoice images/PDFs. Gemini reads the RETAIL qty sold from each sheet, you review, then download the completed Excel.")
    st.warning("⚠️ Gemini can misread handwriting — always review the table before creating the final Excel.")

    if not st.secrets.get("GEMINI_API_KEY", ""):
        st.error("GEMINI_API_KEY is missing. Add it in Streamlit Secrets.")
        st.code('GEMINI_API_KEY = "your_api_key_here"', language="toml")
        return

    master_file = st.file_uploader("Upload master Excel file (.xlsx)", type=["xlsx"], key="ocr-master")
    invoice_files = st.file_uploader(
        "Upload invoice images or PDFs",
        type=["jpg", "jpeg", "png", "heic", "heif", "pdf"],
        accept_multiple_files=True,
        key="ocr-invoices",
    )

    if not master_file or not invoice_files:
        st.info("Upload the master Excel file and all invoice files to begin.")
        return

    if st.button("🔍 Extract invoices with Gemini", type="primary"):
        sheet_entries = {show: {p: 0 for p in PRODUCTS} for show in SHOWS}
        raw_previews = []
        model = get_gemini_model()

        progress = st.progress(0, text="Starting…")
        for i, invoice in enumerate(invoice_files):
            progress.progress((i) / len(invoice_files), text=f"Reading {invoice.name}…")
            try:
                show_name = detect_show_from_filename(invoice.name)
                entries, raw_text = extract_entries_from_invoice(invoice, model)

                if show_name:
                    for p, qty in entries.items():
                        sheet_entries[show_name][p] = sheet_entries[show_name].get(p, 0) + qty

                raw_previews.append({
                    "file": invoice.name,
                    "detected_show": show_name or "⚠️ Not detected",
                    "non_zero_items": {k: v for k, v in entries.items() if v > 0},
                    "raw_text": raw_text,  # full text, not truncated
                })
            except Exception as exc:
                st.error(f"Could not extract {invoice.name}: {exc}")
                raw_previews.append({
                    "file": invoice.name,
                    "detected_show": "❌ Error",
                    "non_zero_items": {},
                    "raw_text": str(exc),
                })
        progress.progress(1.0, text="Done!")

        st.session_state["ocr_sheet_entries"] = sheet_entries
        st.session_state["ocr_raw_previews"] = raw_previews
        st.session_state["ocr_master_bytes"] = master_file.getvalue()
        # Clear stale editor state so fresh Gemini data populates the table
        for key in ["ocr_edited_entries", "ocr_review_editor", "_creating_excel"]:
            st.session_state.pop(key, None)
        st.success(f"✅ Extracted {len(invoice_files)} invoice(s). Review the table below.")

    if "ocr_sheet_entries" not in st.session_state:
        return

    st.subheader("Step 1 — Review & correct extracted quantities")
    st.caption("Each row = one show/day. Edit any wrong values here. These are the RETAIL qty sold written into Column I.")

    # Always keep ocr_edited_entries in sync with ocr_sheet_entries
    # but only initialize it — don't overwrite if user has edited
    if "ocr_edited_entries" not in st.session_state:
        st.session_state["ocr_edited_entries"] = {
            show: dict(prods)
            for show, prods in st.session_state["ocr_sheet_entries"].items()
        }

    review_rows = entries_to_review_rows(st.session_state["ocr_sheet_entries"])

    st.caption("💡 Edit values directly in the table below, then click Create Excel.")
    edited_df = st.data_editor(
        review_rows,
        use_container_width=True,
        num_rows="fixed",
        disabled=["Show"],
        key="ocr_review_editor",
    )

    # Persist whatever is currently in the editor into session state
    # This runs on EVERY render, but we read it back only when button is clicked
    try:
        import pandas as pd
        if isinstance(edited_df, pd.DataFrame):
            edited_list = edited_df.to_dict("records")
        else:
            edited_list = edited_df
        st.session_state["ocr_edited_entries"] = review_rows_to_sheet_entries(edited_list)
    except Exception:
        pass

    with st.expander("🔬 Gemini raw output — check this if values are wrong", expanded=True):
        for preview in st.session_state.get("ocr_raw_previews", []):
            st.markdown(f"**{preview['file']}** → detected show: `{preview['detected_show']}`")
            col1, col2 = st.columns(2)
            with col1:
                st.caption("Non-zero items extracted:")
                st.write(preview["non_zero_items"] or "None")
            with col2:
                st.caption("Raw Gemini response:")
                st.code(preview["raw_text"], language="text")
            st.divider()

    st.subheader("Step 2 — Create Excel from reviewed values")
    if st.button("📥 Create completed Excel", type="primary"):
        try:
            # Read directly from ocr_sheet_entries (the raw Gemini output)
            # because ocr_edited_entries may have been re-zeroed by the rerun
            # Use whichever has more non-zero values
            gemini_entries = st.session_state["ocr_sheet_entries"]
            edited_entries = st.session_state.get("ocr_edited_entries", gemini_entries)

            gemini_total = sum(q for prods in gemini_entries.values() for q in prods.values())
            edited_total = sum(q for prods in edited_entries.values() for q in prods.values())
            corrected = edited_entries if edited_total >= gemini_total else gemini_entries

            non_zero_filtered = {
                show: {p: q for p, q in prods.items() if q > 0}
                for show, prods in corrected.items()
                if any(q > 0 for q in prods.values())
            }
            if non_zero_filtered:
                st.info("Values being written to Excel:")
                st.write(non_zero_filtered)
            else:
                st.warning("⚠️ All quantities are 0 — nothing will be written. Open the Gemini raw output above to see what was extracted.")
                return
            master_bytes = st.session_state["ocr_master_bytes"]
            output, summary = build_download(master_bytes, corrected)
            st.success("Excel ready! Download below.")
            st.download_button(
                "⬇️ Download completed Excel",
                data=output,
                file_name="Fallen_Angels_Completed.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            st.dataframe(summary, use_container_width=True)
        except Exception as exc:
            st.error(f"Could not create Excel: {exc}")
        finally:
            st.session_state.pop("_creating_excel", None)


def fallen_angels_manual_processor():
    st.header("Fallen Angels — Manual Retail Entry")
    st.caption("No AI needed. Enter RETAIL qty sold from each invoice manually, then download the completed Excel.")

    master_file = st.file_uploader("Upload master Excel file (.xlsx)", type=["xlsx"], key="manual-master")
    st.divider()

    sheet_entries = {}
    tabs = st.tabs(SHOWS)
    for show_name, tab in zip(SHOWS, tabs):
        with tab:
            st.subheader(show_name)
            cols = st.columns(3)
            values = {}
            for idx, product in enumerate(PRODUCTS):
                with cols[idx % 3]:
                    values[product] = st.number_input(
                        product, min_value=0, max_value=999, value=0, step=1,
                        key=f"manual-{show_name}-{product}",
                    )
            sheet_entries[show_name] = values

    st.divider()
    if st.button("📥 Create completed Excel", type="primary", disabled=master_file is None):
        try:
            output, summary = build_download(master_file.getvalue(), sheet_entries)
            st.success("Done! Download your completed master file below.")
            st.download_button(
                "⬇️ Download completed Excel",
                data=output,
                file_name="Fallen_Angels_Manual_Completed.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            st.dataframe(summary, use_container_width=True)
        except Exception as exc:
            st.error(f"Could not create the file: {exc}")


def simple_excel_column_updater():
    st.header("Simple Excel Column Updater")
    excel_file = st.file_uploader("Upload Excel file", type=["xlsx"], key="simple-excel")
    sheet_name = st.text_input("Sheet name", value="Sheet1")
    column_letter = st.text_input("Column letter", value="I", max_chars=3)
    start_row = st.number_input("Start row", min_value=1, value=2, step=1)
    end_row = st.number_input("End row", min_value=1, value=10, step=1)
    value = st.text_input("Value to write")
    if st.button("Create updated Excel", disabled=excel_file is None):
        try:
            wb = openpyxl.load_workbook(excel_file)
            if sheet_name not in wb.sheetnames:
                st.error(f"Sheet '{sheet_name}' not found. Available: {', '.join(wb.sheetnames)}")
                return
            ws = wb[sheet_name]
            for row in range(int(start_row), int(end_row) + 1):
                ws[f"{column_letter.upper()}{row}"] = value
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            st.success("Done.")
            st.download_button("⬇️ Download updated Excel", data=output,
                               file_name="Updated_Excel.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        except Exception as exc:
            st.error(f"Error: {exc}")


def heic_to_jpg_converter():
    st.header("HEIC → JPG Converter")
    uploaded_files = st.file_uploader("Upload HEIC/HEIF images", type=["heic", "heif"],
                                      accept_multiple_files=True, key="heic-files")
    quality = st.slider("JPG quality", 60, 100, 90, 5)
    if not uploaded_files:
        st.info("Upload one or more .heic / .heif files.")
        return
    converted = []
    for f in uploaded_files:
        try:
            img = Image.open(f)
            if img.mode not in ("RGB", "L"):
                img = img.convert("RGB")
            buf = io.BytesIO()
            img.save(buf, format="JPEG", quality=quality, optimize=True)
            buf.seek(0)
            jpg_name = f.name.rsplit(".", 1)[0] + ".jpg"
            converted.append((jpg_name, buf.getvalue()))
            st.success(f"Converted: {f.name} → {jpg_name}")
            st.image(buf.getvalue(), caption=jpg_name, use_container_width=True)
            st.download_button(f"Download {jpg_name}", data=buf.getvalue(),
                               file_name=jpg_name, mime="image/jpeg",
                               key=f"dl-{jpg_name}")
        except Exception as exc:
            st.error(f"Could not convert {f.name}: {exc}")
    if len(converted) > 1:
        zip_buf = io.BytesIO()
        with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
            for name, data in converted:
                zf.writestr(name, data)
        zip_buf.seek(0)
        st.download_button("⬇️ Download all as ZIP", data=zip_buf,
                           file_name="converted_jpgs.zip", mime="application/zip")


# ─── Main ────────────────────────────────────────────────────────────────────

def main():
    if not check_password():
        return
    logout_button()
    st.title("Niraj Excel Tools")
    st.caption("Automated OCR (Gemini) • Manual entry • HEIC converter")

    task = st.sidebar.selectbox("Choose task", [
        "Automated invoice OCR → Excel",
        "Manual retail entry → Excel",
        "Simple Excel column updater",
        "HEIC → JPG Converter",
    ])

    st.sidebar.info("OCR tool uses Gemini AI. Manual entry and HEIC converter work offline.")

    if task == "Automated invoice OCR → Excel":
        automated_ocr_processor()
    elif task == "Manual retail entry → Excel":
        fallen_angels_manual_processor()
    elif task == "Simple Excel column updater":
        simple_excel_column_updater()
    elif task == "HEIC → JPG Converter":
        heic_to_jpg_converter()


if __name__ == "__main__":
    main()
